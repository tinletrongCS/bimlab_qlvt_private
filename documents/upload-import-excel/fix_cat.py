import openpyxl

excel_file = r'D:\BIMLAB\bimlab_qlvt\documents\upload-import-excel\mau_import_danh_sach_tai_san_bimlab_updated.xlsx'
wb = openpyxl.load_workbook(excel_file)
ws = wb['HoSoTaiSan_import']

for r_idx in range(46, 50):
    pl = ws.cell(row=r_idx, column=3).value
    danh_muc = 'TSCD_DESKTOP_COMPUTER' if pl == 'FIXED_ASSET' else 'CCDC_DESKTOP_COMPUTER'
    ws.cell(row=r_idx, column=5).value = danh_muc

wb.save(excel_file)
print('Fixed last 4 rows.')
