import pandas as pd
import numpy as np
import math

csv_file = r'D:\BIMLAB\bimlab_qlvt\documents\upload-import-excel\Danh sách sản phẩm mẫu(Sheet1).csv'
excel_file = r'D:\BIMLAB\bimlab_qlvt\documents\upload-import-excel\mau_import_danh_sach_tai_san_bimlab.xlsx'
out_file = r'D:\BIMLAB\bimlab_qlvt\documents\upload-import-excel\mau_import_danh_sach_tai_san_bimlab_updated.xlsx'

try:
    df_csv = pd.read_csv(csv_file, encoding='utf-8-sig', encoding_errors='replace')
    if 'STT' not in df_csv.columns[0]:
        df_csv = pd.read_csv(csv_file, encoding='cp1258', encoding_errors='replace')
except:
    df_csv = pd.read_csv(csv_file, encoding='cp1258', encoding_errors='replace')

rows = []
for idx in range(len(df_csv)):
    row = df_csv.iloc[idx]
    
    name = str(row.iloc[1]).strip()
    if name == 'nan' or name == '':
        continue
        
    sl = str(row.iloc[2]).strip()
    try:
        sl = int(sl)
    except:
        sl = 1
        
    price = str(row.iloc[3]).replace(',', '').strip()
    try:
        price = float(price)
    except:
        price = 0
        
    pl = str(row.iloc[6]).strip()
    nct = str(row.iloc[7]).strip()
    note = str(row.iloc[8]).strip()
    if note == 'nan':
        note = ''
    
    phan_loai = 'FIXED_ASSET' if 'TSC' in pl else 'TOOL_EQUIPMENT'
    
    phan_loai_lop_con = ''
    if phan_loai == 'FIXED_ASSET':
        phan_loai_lop_con = 'TANGIBLE' if ('h' in nct.lower() and 'u' in nct.lower()) else 'INTANGIBLE'
    else:
        phan_loai_lop_con = 'MULTI_USE' if 'nhi' in nct.lower() else 'SINGLE_USE'
        
    danh_muc = ''
    n_lower = name.lower()
    
    if 'máy tính' in n_lower or 'máy trạm' in n_lower or 'desktop' in n_lower or 'bàn' in n_lower:
        danh_muc = 'TSCD_DESKTOP_COMPUTER' if phan_loai == 'FIXED_ASSET' else 'CCDC_DESKTOP_COMPUTER'
    if 'laptop' in n_lower:
        danh_muc = 'TSCD_LAPTOP' if phan_loai == 'FIXED_ASSET' else 'CCDC_LAPTOP'
    if 'màn hình' in n_lower:
        danh_muc = 'TSCD_MONITOR' if phan_loai == 'FIXED_ASSET' else 'CCDC_MONITOR'
    if 'máy chủ' in n_lower or ('server' in n_lower and 'windows' not in n_lower and 'sql' not in n_lower and 'security' not in n_lower and 'backup' not in n_lower):
        danh_muc = 'TSCD_SERVER'
    if 'lưu trữ' in n_lower or 'backup' in n_lower and 'cloud' not in n_lower:
        danh_muc = 'TSCD_STORAGE_DEVICE' if phan_loai == 'FIXED_ASSET' else 'CCDC_STORAGE_DEVICE'
    if 'phần mềm' in note.lower() or 'windows' in n_lower or 'kaspersky' in n_lower or 'office' in n_lower or 'claude' in n_lower or 'cloud' in n_lower or 'license' in note.lower() or 'tài khoản' in note.lower() or 'thuê bao' in note.lower():
        danh_muc = 'TSCD_SOFTWARE_LICENSE' if phan_loai == 'FIXED_ASSET' else 'CCDC_SOFTWARE'
        if phan_loai == 'FIXED_ASSET':
            phan_loai_lop_con = 'INTANGIBLE'

    new_row = {
        'Mã tài sản': '',
        'Tên tài sản*': name,
        'Phân loại*': phan_loai,
        'Phân loại lớp con*': phan_loai_lop_con,
        'Danh mục tài sản*': danh_muc,
        'Phòng ban': '',
        'Chi nhánh': '',
        'Mẫu tài sản': '',
        'Phương pháp khấu hao': '',
        'Số Series/MAC': '',
        'Ngày bắt đầu khấu hao': '',
        'Ngày sử dụng': '',
        'Số tháng': '',
        'Nguyên giá tài sản': price,
        'Giá trị sổ sách': '',
        'Trạng thái tài sản': '',
        'Mã quốc gia/Xuất xứ': '',
        'Năm sản xuất': '',
        'Năm lắp đặt/Cài đặt': '',
        'Mô tả kỹ thuật': note
    }
        
    for i in range(sl):
        rows.append(new_row.copy())

df_new = pd.DataFrame(rows)

import openpyxl
wb = openpyxl.load_workbook(excel_file)
ws = wb['HoSoTaiSan_import']
for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
    for cell in row:
        cell.value = None

from openpyxl.utils.dataframe import dataframe_to_rows
for r_idx, row in enumerate(dataframe_to_rows(df_new, index=False, header=False), 4):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(out_file)
print('Mapped {} items into {} rows in HoSoTaiSan_import. Saved to _updated file.'.format(len(df_csv), len(df_new)))
